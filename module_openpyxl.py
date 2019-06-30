#!/usr/bin/env python3

# Import the open python xl module, as xl.
import openpyxl as xl

# From xl use the load_workbook method and pass in the excel filename that is
# in the current working dir, and assign it to var obj 'wb'.
wb = xl.load_workbook("transactions.xlsx")
sheet = wb["Sheet1"]    # Because this xls only has one sheet, the name of the
                        # sheet must be set in the list and it is case-sensitive.
cell = sheet["a1"]      # Returns the value of the first cell and assign it to
                        # var obj 'cell'.
cell = sheet.cell(1, 1) # This line returns the exact same value as line 11.
print(cell.value)


